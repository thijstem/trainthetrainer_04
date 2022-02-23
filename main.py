#######################################################################################################################
# A. Importeer csv (uit adlib) naar dataframe (pandas)

# pip install pandas
import pandas as pd

# inlezen csv
# ingeven padnaam naar csv, en indien van toepassing aanvullen delimiter
lijstDMG = pd.read_csv(r"export (16).csv", delimiter=';')

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 1. Selecteer ontbrekende velden (via pandas)
# 1.1 Objectnaam

# identificeer wanneer objectnaam 'leeg' is
ontbrekende_objectnaam = pd.isna(lijstDMG['objectnaam'])

# hou vanuit de lijst enkel de records over waarbij objectnaam leeg = true was
ontbrekende_objectnaam = lijstDMG[ontbrekende_objectnaam]

# 1.2 Titel

ontbrekende_titel = pd.isna(lijstDMG['titel'])
ontbrekende_titel = lijstDMG[ontbrekende_titel]

# 1.3 Beschrijving

ontbrekende_beschrijving = pd.isna(lijstDMG['beschrijving'])
ontbrekende_beschrijving = lijstDMG[ontbrekende_beschrijving]

# 1.4 Vervaardiger

ontbrekende_vervaardiger = pd.isna(lijstDMG['vervaardiger'])
ontbrekende_vervaardiger = lijstDMG[ontbrekende_vervaardiger]

# 1.5 Datering_begin

ontbrekende_datering = pd.isna(lijstDMG['vervaardiging.datum.begin'])
ontbrekende_datering = lijstDMG[ontbrekende_datering]

# 1.6 Datering_eind

ontbrekende_datering_eind = pd.isna(lijstDMG['vervaardiging.datum.eind'])
ontbrekende_datering_eind = lijstDMG[ontbrekende_datering_eind]

# 1.7 Instelling

ontbrekende_instelling = pd.isna(lijstDMG['instelling.naam'])
ontbrekende_instelling = lijstDMG[ontbrekende_instelling]

# 1.8 Plaats

ontbrekende_vervaardiging_plaats = pd.isna(lijstDMG['vervaardiging.plaats'])
ontbrekende_vervaardiging_plaats = lijstDMG[ontbrekende_vervaardiging_plaats]

# 1.9 Rol

ontbrekende_vervaardiger_rol = pd.isna(lijstDMG['vervaardiger.rol'])
ontbrekende_vervaardiger_rol= lijstDMG[ontbrekende_vervaardiger_rol]

# 1.10 Rechten.type

ontbrekende_rechtentype = pd.isna(lijstDMG['rechten.type'])
ontbrekende_rechtentype = lijstDMG[ontbrekende_rechtentype]

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 2. Output (excel) maken met lijsten met ontbrekende velden (via openpyxl)

# pip install openpyxl
from openpyxl import Workbook

# maak een excel aan
wb = Workbook()

# voeg een sheet toe met ontbrekende velden
# 2.1 objectnaam

from openpyxl.utils.dataframe import dataframe_to_rows

# maak sheet (tabblad)
ws = wb.create_sheet("Objectnaam")
# zet dataframe (pandas) om naar rijen in het tabblad
rows = dataframe_to_rows(ontbrekende_objectnaam, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# sla excel op
wb.save(r"output.xlsx")

# 2.2 titel

ws = wb.create_sheet("Titel")
rows = dataframe_to_rows(ontbrekende_titel, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.3 beschrijving

ws = wb.create_sheet("Beschrijving")
rows = dataframe_to_rows(ontbrekende_beschrijving, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.4 vervaardiger

ws = wb.create_sheet("Vervaardiger")
rows = dataframe_to_rows(ontbrekende_vervaardiger, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.5 datering_begin

ws = wb.create_sheet("Datering_begin")
rows = dataframe_to_rows(ontbrekende_datering, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.6 datering_eind

ws = wb.create_sheet("Datering_eind")
rows = dataframe_to_rows(ontbrekende_datering_eind, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.7 Plaats

ws = wb.create_sheet("Ontbrekende plaats")
rows = dataframe_to_rows(ontbrekende_vervaardiging_plaats, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.8 Rol vervaardiger

ws = wb.create_sheet("Ontbrekende rol vervaardiger")
rows = dataframe_to_rows(ontbrekende_vervaardiger_rol, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# 2.9 Rechten.type

ws = wb.create_sheet("Ontbrekende rechtentype")
rows = dataframe_to_rows(ontbrekende_rechtentype, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

wb.save(r"output.xlsx")

# C. visualisatie van de data
# 1. tellen van de data (pandas)

# tel aantal keer objectnummer aanwezig is in lijst records zonder objectnamen (gezien objectnummer altijd aanwezig)
aantal_ontbrekende_objectnamen = ontbrekende_objectnaam['objectnummer'].count()
print(aantal_ontbrekende_objectnamen)
# doe hetzelfde voor de overige velden
aantal_ontbrekende_titels = ontbrekende_titel['objectnummer'].count()
aantal_ontbrekende_beschrijvingen = ontbrekende_beschrijving['objectnummer'].count()
aantal_ontbrekende_vervaardigers = ontbrekende_vervaardiger['objectnummer'].count()
aantal_ontbrekende_dateringbegin = ontbrekende_datering['objectnummer'].count()
aantal_ontbrekende_plaatsen = ontbrekende_vervaardiging_plaats['objectnummer'].count()
aantal_ontbrekende_rechtentypen = ontbrekende_rechtentype['objectnummer'].count()


# C. visualisatie van de data
# 2. weergeven van de data in grafiek in excel (openpyxl)

# 2.1 voeg de data toe in excel
# zet de data in een list
labels = ["objectnaam", "titel", "beschrijving", "vervaardiger", "datering", "plaats", "rechtentype"]
ontbrekende_data = [aantal_ontbrekende_objectnamen, aantal_ontbrekende_titels, aantal_ontbrekende_beschrijvingen,
                    aantal_ontbrekende_vervaardigers, aantal_ontbrekende_dateringbegin, aantal_ontbrekende_plaatsen, aantal_ontbrekende_rechtentypen]

# voeg de list toe aan de excel (openpyxl)

ws = wb.active
ws.title = 'Info'
# titel lijst
ws['A1'] = "Ontbrekende velden"
wb.save(r"output.xlsx")

# de labels
rij1 = 2
for label in labels:
    ws.cell(row=rij1, column=1).value = label
    rij1 += 1
wb.save(r"output.xlsx")

# de waardes
rij2 = 2
for veld in ontbrekende_data:
    ws.cell(row=rij2, column=2).value = veld
    rij2 += 1
wb.save(r"output.xlsx")

# 2.2 maak de grafiek in excel

from openpyxl.chart import BarChart3D, Reference

# selecteer de data
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=8)

# selecteer de labels
labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=8)

# creeer grafiek
chart = BarChart3D()

# voeg grafiek toe aan excel
ws.add_chart(chart, "E2")
wb.save(r"output.xlsx")

# voeg titel toe
chart.title = 'Ontbrekende data'

# voeg namen x en y as toe
chart.y_axis.title = 'aantal'
chart.x_axis.title = 'velden'

# voeg de data & labels toe
chart.add_data(data)
chart.set_categories(labels)
chart.style = 14

# sla excel op
wb.save(r"output.xlsx")