# Script om een adlibexport.csv bestand te controleren en alle records te returnen met ontbrekende waarden
# Een snelle manier om ontbrekende data te identificeren en daarna in adlib/axiell aan te vullen


import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, BarChart3D, Reference
from openpyxl.chart.label import DataLabelList


# NODIG: een adlib-export met de kolommen die je wil controleren
# lees de csv in: BESTANDSNAAM AANPASSEN en sla op in een dataframe
df = pd.read_csv(r"export (23).csv", delimiter=';')

# maak een lijst van alle kolomnamen (gekozen adlibvelden in export)
velden = list(df.columns)

# maak een nieuw workbook
wb = Workbook()

# voor elke kolom: maakt worksheet aan met "geen x" als titel + maakt dataframe met alle records met ontbrekende waarde
# + bewaar die lijst en print die als rijen in het aangemaakte worksheet + sla de creatie op in de projectmap
for i in velden:
    ws = wb.create_sheet("geen " + str(i))
    i = pd.isna(df[i])
    i = df[i]
    rows = dataframe_to_rows(i, index=False)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    wb.save(r"export28.xlsx")

wb = load_workbook("export28.xlsx")
ws = wb.active
ws.title = "Info"
labels = wb.sheetnames
ws["A1"] = "Ontbrekende velden"
ws["A2"] = "Aantal"
rij1 = 2
for label in labels:
    ws.cell(row=rij1, column=1).value = label
    rij1 += 1
rij2 = 2
for label in labels:
    ws = wb[label]
    aantal = ws.max_row - 1
    ws = wb["Info"]
    ws.cell(row=rij2, column=2).value = aantal
    rij2 += 1

ws.delete_rows(2)
data = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=ws.max_row)

# selecteer de labels
labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=ws.max_row)

# creeer grafiek
chart = BarChart()
chart.height = 10
chart.width = 20

# voeg grafiek toe aan excel
ws.add_chart(chart, "E2")

chart.title = 'Ontbrekende data'

# voeg namen x en y as toe
chart.y_axis.title = 'Aantal'
chart.x_axis.title = 'Velden'

# voeg de data & labels toe
chart.add_data(data)
chart.set_categories(labels)
chart.dataLabels = DataLabelList()
chart.dataLabels.showVal = True

wb.save(r"export28_met_aantal.xlsx")