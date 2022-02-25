#######################################################################################################################
# A. Importeer csv (uit adlib) naar dataframe (pandas)

# pip install pandas
import pandas as pd

# inlezen csv
# ingeven padnaam naar csv, en indien van toepassing aanvullen delimiter
lijstDMG = pd.read_csv(r"rechtenvelden.csv", delimiter=';')

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 1. Selecteer ontbrekende velden (via pandas)
# 1.1 Objectnaam

# identificeer wanneer objectnaam 'leeg' is

# hou vanuit de lijst enkel de records over waarbij objectnaam leeg = true was


# 1.3 Met rechtentype

met_rechtentype = pd.notna(lijstDMG['rechten.type'])
met_rechtentype = lijstDMG[met_rechtentype]

#######################################################################################################################
# B. Maak lijsten met ontbrekende velden
# 2. Output (excel) maken met lijsten met ontbrekende velden (via openpyxl)

# pip install openpyxl
from openpyxl import Workbook

# maak een excel aan
wb = Workbook()

# voeg een sheet toe met ontbrekende velden
# 2.1 objectnaam

from openpyxl.utils.dataframe import dataframe_to_rows

# maak sheet (tabblad)
ws = wb.create_sheet("MetRechtentype")
# zet dataframe (pandas) om naar rijen in het tabblad
rows = dataframe_to_rows(met_rechtentype, index=False)
for r_idx, row in enumerate(rows, 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# sla excel op
wb.save(r"metrechtentype.xlsx")
